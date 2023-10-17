import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

# wb = openpyxl.load_workbook('test.xlsx')
# sheet = wb['Feuille 1']
# cell = sheet['A1']
# #ou cell = sheet.cell(row = 1, column = 1)
# print(cell.value)
# print(cell.coordinate)
# print(get_column_letter(30))
# print(column_index_from_string('E'))

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb['Feuille 1']
cell = sheet['A1']

# sheet.title = 'Inventaire fruits'
# cell.value = 'Kiwis'

# wb.save('test_copy.xlsx')

cell = sheet['A2']
cell.value = 'Pommes de rainette'
sheet.column_dimensions['A'].width = 30
sheet.row_dimensions[2].height = 30
cell.font = Font(size = 15, bold = True)
cell.alignment = Alignment(horizontal = "center", vertical = "center")
cell.fill = PatternFill("solid", start_color = "0ebf3e")
cell.border = Border(left = Side(style='thick'), right = Side(style='thick'), top = Side(style='thick'), bottom = Side(style='thick'))

sheet.merge_cells('A3:A4')

wb.save('test_copy.xlsx')