import openpyxl
from openpyxl.styles import Font, PatternFill
from create_timetable import Create_timetable


# Vert : #93c47d
# Bleu : #6d9eeb
# Saumon : #ea9999

# wb = openpyxl.load_workbook('edt Do_It.23-24.xlsx')
# sheet = wb['année']
# cell = sheet.cell(row = 6, column = 22)
# # cell = sheet['V6']
# word = ''
# for char in cell.value:
#     word += char
#     if "\n" in word:
#         word = word[:len(word) - 1]
#         break
# print(word)
# test = Create_timetable()
# cell = test.sheet2['V6']
# print(test.get_course_duration(cell))
# # font_test = Font(size = 24, bold = True)

# # cell2 = sheet.cell(row = 12, column = 3)
# # sheet.unmerge_cells('AL9:AO9')
test = Create_timetable()
# test.create_timetable_automatic("Dang vu", "duc")
# test.remove_course("edt de Duc Dang Vu.xlsx")
test.create_all_timetable()
# test.remove_course("edt Do_It.23-24.xlsx")


