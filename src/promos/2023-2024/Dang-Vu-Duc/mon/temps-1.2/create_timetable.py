import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class Create_timetable():
    def __init__(self):
        self.create_new_sheets()
        self.weeks = []
        self.possible_times = [8, 9, 10, 11, 12, 14, 15, 16, 17, 18]
        self.wb = openpyxl.load_workbook('edt Do_It.23-24.xlsx')
        self.sheet = self.wb['année']
        for i in range(3,35):
            cell_begin = self.sheet.cell(row = i, column = 3)
            cell_end = self.sheet.cell(row = i, column = 4)
            self.weeks.append([cell_begin.value, cell_end.value])
    
    def create_new_sheets(self):
        self.wb2 = openpyxl.load_workbook('étudiants_Do-It_23_24.xlsx')
        self.sheet2 = self.wb2['effectif']
        self.wb3 = openpyxl.load_workbook('edt Do_It.23-24.xlsx')
        self.sheet3 = self.wb3['année']
    
    def get_cell_reference(self, day, month, time):
        nb_week = None
        nb_day = None
        for k in range(len(self.weeks)):
            date_begin = self.weeks[k][0]
            date_end = self.weeks[k][1]
            if date_begin.month == date_end.month and month == date_begin.month:
                if day in range(date_begin.day, date_end.day + 1):
                    nb_week = k
                    nb_day = day - date_begin.day
                    break
            if date_begin.month != date_end.month and month == date_begin.month:
                if day >= date_begin.day:
                    nb_week = k
                    nb_day = day - date_begin.day
                    break
            if date_begin.month != date_end.month and month == date_end.month:
                if day <= date_end.day:
                    nb_week = k
                    virtual_day_begin = date_end.day - 4
                    nb_day = day - virtual_day_begin
                    break
        if nb_week == None:
            print("La date indiquée n'est pas valide")
            return(False)
        nb_row = nb_week + 3
        if time not in self.possible_times:
            print("L'heure indiquée n'est pas valide")
            return(False)
        else:
            index = self.possible_times.index(time)
            nb_column = nb_day * 8 + 5 + index
        return(get_column_letter(nb_column) + str(nb_row))

    def add_course(self, file_name):
        current_wb = openpyxl.load_workbook(file_name)
        current_sheet = current_wb['année']
        day = int(input("Entrez le jour du cours: "))
        month = int(input("Entrez le mois du cours: "))
        time = None
        while time not in self.possible_times:
            time = int(input("Entrez l'heure du cours: "))
        duration = int(input("Entrez la durée du cours: "))
        name = str(input("Entrez le nom du cours: "))
        professor = str(input("Entrez le nom du professeur: "))
        room = str(input("Entrez le numéro de salle: "))
        color = str(input("Entrez le code hex de la couleur du cours: "))
        cell_reference = self.get_cell_reference(day, month, time)
        cell_reference2 = self.get_cell_reference(day, month, time + duration - 1)
        current_sheet.merge_cells(cell_reference + ':' + cell_reference2)
        cell = current_sheet[cell_reference]
        cell.fill = PatternFill("solid", start_color = color)
        cell.value = f"{name}\n{professor}\n{room}"
        cell.alignment = Alignment(wrapText=True, horizontal = "center", vertical = "center")
        current_sheet.row_dimensions[cell.row].height = 10 * 3 + 10
        current_wb.save(file_name)
    
    def remove_course(self, file_name):
        current_wb = openpyxl.load_workbook(file_name)
        current_sheet = current_wb['année']
        day = int(input("Entrez le jour du cours qui doit être supprimé: "))
        month = int(input("Entrez le mois du cours qui doit être supprimé: "))
        time = int(input("Entrez l'heure du cours qui doit être supprimé: "))
        duration = int(input("Entrez la durée du cours qui doit être supprimé: "))
        cell_reference = self.get_cell_reference(day, month, time)
        cell_reference2 = self.get_cell_reference(day, month, time + duration - 1)
        border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
        current_sheet.unmerge_cells(cell_reference + ':' + cell_reference2)
        for k in range(duration):
            cell_reference = self.get_cell_reference(day, month, time + k)
            cell = current_sheet[cell_reference]
            cell.border = border
            cell.fill = PatternFill("solid", start_color = "FFFFFF")
            cell.value = None
        current_wb.save(file_name)
    
    def remove_course_automatic(self, cell, duration):
        border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
        if duration == 1:
            cell.border = border
            cell.fill = PatternFill("solid", start_color = "FFFFFF")
            cell.value = None
        else:
            cell_end_coordinate = str(get_column_letter(cell.column + duration - 1)) + str(cell.row)
            self.sheet3.unmerge_cells(cell.coordinate + ':' + cell_end_coordinate)
            for k in range(duration):
                current_cell_coordinate = str(get_column_letter(cell.column + k)) + str(cell.row)
                current_cell = self.sheet3[current_cell_coordinate]
                current_cell.border = border
                current_cell.fill = PatternFill("solid", start_color = "FFFFFF")
                current_cell.value = None

    def get_course_duration(self, cell):
        for mergedCell in self.sheet3.merged_cells.ranges:
            if cell.coordinate in mergedCell:
                return(mergedCell.max_col - mergedCell.min_col + 1)
        return(1)
        


    def create_timetable_automatic(self, nom, prénom):
        courses_name = ["bonjour !", "tc1 : agilité", "lancement projet 3a", "tc1 : gestion des sources", "tc1 : service design", "langues", "projet 3a", "vacances", "filière métier", "tronc commun 3a", "prez mon 2", "prez pok", "point pok sprint 1", "point pok sprint 2", "prez mon 1", "pok&mon", "cap 1a/3a/conception si", "prez projet", "rencontre w3g", "débrief", "conférence métier"]
        row_number = None
        for k in range(3,27):
            cell_nom = self.sheet2.cell(row = k, column = 1)
            cell_prénom = self.sheet2.cell(row = k, column = 2)
            if prénom.lower() == cell_prénom.value.lower() and nom.lower() == cell_nom.value.lower():
                row_number = k
                break
        if row_number == None:
            print("Nom ou prénom invalide")
            return()
        for k in range(7,30):
            cell = self.sheet2.cell(row = row_number, column = k)
            if cell.value == 'x' or cell.value == "X":
                course = self.sheet2.cell(row = 2, column = k).value
                course = course.lower()
                if course == 'bonnes pratiques':
                    courses_name.append("structuration d'un projet info")
                    courses_name.append("programmation par les tests")
                else:
                    courses_name.append(course)
        border_medium_left =  Border(left = Side(style='medium'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
        border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
        border_thick_left = Border(left = Side(style='thick'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
        for nb_row in range(5, 35):
            for nb_column in range(5, 45):
                cell = self.sheet3.cell(row = nb_row, column = nb_column)
                course_name = ''
                if cell.value != None:
                    for char in cell.value:
                        course_name += char.lower()
                        if "\n" in course_name:
                            course_name = course_name[:len(course_name) - 1]
                            break
                    if course_name not in courses_name:
                        course_duration = self.get_course_duration(cell)
                        self.remove_course_automatic(cell, course_duration)
                    else:
                        cell.alignment = Alignment(wrapText=True, horizontal = "center", vertical = "center")
                        cell.font = Font(size = 10)
                        self.sheet3.row_dimensions[cell.row].height = 10 * 3 + 10
                if (nb_column - 5) % 8 == 0:
                    cell.border = border_thick_left
                elif (nb_column - 5) % 8 == 4:
                    cell.border = border_medium_left
                else:
                    cell.border = border
        self.wb3.save(f'edt de {cell_prénom.value} {cell_nom.value}.xlsx')
        print(f'Emploi du temps enregistré sous le nom de "edt de {cell_prénom.value} {cell_nom.value}.xlsx"')

    def create_all_timetable(self):
        for k in range(3,27):
            self.create_new_sheets()
            nom = self.sheet2.cell(row = k, column = 1).value
            prénom = self.sheet2.cell(row = k, column = 2).value
            self.create_timetable_automatic(nom, prénom)
        print("Tout les edt ont été enregistrés")